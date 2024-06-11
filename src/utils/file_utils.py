import os
from typing import Optional
import warnings
warnings.simplefilter(action="ignore", category=FutureWarning)
import pandas as pd
warnings.simplefilter(action="ignore", category=pd.errors.ParserWarning)
warnings.simplefilter(action="ignore", category=UserWarning)
from location import LocationPoint
from sign_plan import SignPlan

def add_coloring(df: pd.DataFrame, manual_rows: set[int], column: str = "OPERAATIO"):

    """
    Adds coloring to dataframe column to all rows indexed by notify rows.
    """

    def highlight_rows(row: pd.Series):
        if row.name in manual_rows:
            return  len(row)* ["background-color: yellow"]
        else:
            return len(row)*[""]
    style = df.style.apply(highlight_rows,axis=1,subset=column)

    return style


def create_dir_if_not_exist(directory: str):

    """
    Creates directory with given name in the current working directory.
    :return: Path to newly created directory.
    """

    path = os.path.join(os.getcwd(),directory)
    if not os.path.exists(path):
        os.makedirs(path)

    return path

def preprocess_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Removes possible line-endings and ' characters from dataframe.
    """
    df = df.replace(r'\n', ' ', regex=True)
    df = df.map(lambda x: str(x).replace(r"'",""),na_action="ignore")

    return df

def preprocess_signplan(df: pd.DataFrame) -> pd.DataFrame:

    """
    Applies preprocessing to signplan dataframe.
    :return: Processed dataframe.
    """

    # Simplify and unify naming to ease the processing.
    df.rename(columns =
              {
                  "MERKKI / MERKINTÄ": "MERKKI/MERKINTÄ",
                  "VALMISTUSNUMERO": "MERKIN VALMISTUSNUMERO",
                  "KIINNITYSPISTE": "KIINNITYS",
                  "TULPPA ": "TULPPA",
                  "SIJAINTIRAIDE": "RAIDE",
                  "KIINNIKKEET": "KIINNITYSTARVIKKEET",
                  "LKP/ LKPVÄLI": "LKP/LKPVÄLI",
                  "LKP /  LKPVÄLI": "LKP/LKPVÄLI",
                  "ASENNUSETÄISYYS (mm)": "ASENNUSETÄISYYS",
                  "ASENNUSKORKEUS (mm)": "ASENNUSKORKEUS"},
              inplace=True)
    
    # Certain signs are skipped completely.
    df = df[~(df["MERKKI/MERKINTÄ"].str.contains("(?i)vanha aurausmerkki|vanha hengenvaaramerkki",na=False))].reset_index(drop=True)

    # Remove unnamed columns caused by the formatting used in excel files.
    df.drop(df.columns[df.columns.str.startswith("Unnamed")],axis=1,inplace=True)

    # Old files use "Kasvava" instead of "Nouseva". Use new terminology.
    df["LUKUSUUNTA"] = df["LUKUSUUNTA"].map(lambda x: x if x.lower() != "kasvava" else "Nouseva",na_action="ignore")

    # Remove whitespaces
    df = df.map(lambda x: x.strip() if isinstance(x,str) else x, na_action="ignore")

    return df

def is_remove_operation(row: pd.Series) -> bool:
    """ Checks if the sign is to be removed. """

    return row["TOIMENPIDE"].lower() == "merkki poistetaan" 

def try_get_sheetname(path: str):

    """
    Exact name varies but the sheet containing actual sign information usually has suunnitelma in its name. 
    """

    from openpyxl import load_workbook
     
    wb = load_workbook(path)
    sheets = wb.sheetnames

    for sheetname in sheets:
        if "suunnitelma" in sheetname.lower():
            return sheetname

    return None

def try_read_excel(fullpath: str) -> Optional[pd.DataFrame]:
    
    sheetname = try_get_sheetname(fullpath)

    if sheetname is None:
        print(f"Merkkisuunnitelmasta {fullpath.split('.')[0]} ei löytynyt merkkisuunnitelma välilehteä.")
        return None
    
    # In some cases track kilometer is not represented in standard way of ####+####
    # in which case the from_str method fails.
    # We dont want to fail the reading of a whole file for a few invalid rows so
    # those rows will be left as is and will be marked to be manually handled later on.

    def read_track_kilometer(x):
        try:
            return str(LocationPoint.from_str(x))
        except:
            return x

    df = pd.read_excel(fullpath, sheet_name=sheetname,dtype=str,
                        converters={"RATAKILOMETRI": read_track_kilometer})
    
    # In some of the files there are additional info rows before the actual content starts
    # in which case the first row is not the header.
    if "TILIRATAOSA" not in df.columns.tolist():
        # Find the first data row, use one of the column names as help.
        bool_array = df == "TILIRATAOSA"
        rows = list(bool_array.any(axis=1))
        idx = next(idx for idx,value in enumerate(rows) if value) # Row-idx of header row
        
        # Reread the file with right position
        df = pd.read_excel(fullpath, sheet_name=sheetname,skiprows=idx+1,dtype=str,
                            converters={"RATAKILOMETRI": read_track_kilometer})

    return df

def directory_traverse(path: str, signplan_paths: list):

    ratko_file_required = False

    plans = {
        "signplans": [],
        "ratko": None
        }
    for name in os.listdir(path):

        fullpath = os.path.join(path,name)
        if os.path.isdir(fullpath):
            directory_traverse(fullpath,signplan_paths)
        else:
            if name.startswith("~"): continue # Skip temporary files

            extension = name[name.rfind(".")+1:] 
            if extension not in ["xlsx","xlsm","csv","ods"]:
                continue
            
            if "suunnitelma" in name.lower() or "sijoitustaulukko" in name.lower():
                print(f"Luetaan tiedostoa: {name}")

                try:
                    df = try_read_excel(fullpath)
                except Exception as err:
                    print(f"Merkkisuunnitelman {name} luku epäonnistui:")
                    print(str(err))
                    print()
                    continue

                if df is None:
                    continue
                
                # Apply signplan specific preprocessing
                df = preprocess_signplan(df)

                # Generic dataframe preprocessing used both for signplans and ratko files.
                df = preprocess_dataframe(df)

                plans["signplans"].append(SignPlan(df,name.split(".")[0]))            
                ratko_file_required = True
        
            elif "ratko" in name or "radan_merkki" in name:
                if plans["ratko"]: # Already exists
                    raise FileExistsError("Merkkisuunnitelma kansiossa on enemmäin kuin yksi ratko-tiedosto.")
                ratko_data = pd.read_csv(fullpath,sep=";",header = [0,1])
                ratko_data = preprocess_dataframe(ratko_data)

                plans["ratko"] = ratko_data

    if ratko_file_required and plans["ratko"] is None:
        print(f"Kohteesta {path} ei löytynyt ratko tiedostoa.")
        return

    if len(plans["signplans"]) > 0:
        signplan_paths.append(plans)

def read_files(directory: str):

    if not os.path.isdir(directory):
        raise ValueError(f"{directory} is not a directory")
    files = []
    directory_traverse(directory,files)

    return files

